"""
NJ High School Sports Article Scraper
Runs daily — fetches NJ.com RSS feed, deduplicates, appends new articles to Excel.
"""

import urllib.request, urllib.parse, xml.etree.ElementTree as ET
import json, re, time, html, os
from datetime import datetime, timezone
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
EXCEL_FILE   = SCRIPT_DIR / "NJ_HS_Sports_Articles_2026.xlsx"
TRACKING_FILE= SCRIPT_DIR / "scraped_urls.json"
LOG_FILE     = SCRIPT_DIR / "scraper_log.txt"
START_DATE   = datetime(2026, 1, 1, tzinfo=timezone.utc)
END_DATE     = datetime(2026, 12, 31, tzinfo=timezone.utc)  # full year going forward
RSS_BASE     = "https://www.nj.com/arc/outboundfeeds/rss/category/highschoolsports/?from={}"
HEADERS      = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

# ── Helpers ─────────────────────────────────────────────────────────────────
def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")

def fetch(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=20) as r:
        return r.read().decode("utf-8")

def parse_sport(title):
    t = title.lower()
    for sport, kws in [
        ("Flag Football",    ["flag football"]),
        ("Football",         [" football"]),
        ("Boys Basketball",  ["boys basketball"]),
        ("Girls Basketball", ["girls basketball"]),
        ("Basketball",       ["basketball"]),
        ("Baseball",         ["baseball", "diamond classic"]),
        ("Softball",         ["softball"]),
        ("Boys Lacrosse",    ["boys lacrosse"]),
        ("Girls Lacrosse",   ["girls lacrosse"]),
        ("Lacrosse",         ["lacrosse"]),
        ("Boys Soccer",      ["boys soccer"]),
        ("Girls Soccer",     ["girls soccer"]),
        ("Soccer",           ["soccer"]),
        ("Wrestling",        ["wrestling"]),
        ("Boys Tennis",      ["boys tennis"]),
        ("Girls Tennis",     ["girls tennis"]),
        ("Tennis",           ["tennis"]),
        ("Track & Field",    ["track & field", "track and field", "track team", "sectional"]),
        ("Boys Volleyball",  ["boys volleyball"]),
        ("Girls Volleyball", ["girls volleyball"]),
        ("Volleyball",       ["volleyball"]),
        ("Swimming",         ["swimming"]),
        ("Cross Country",    ["cross country"]),
        ("Golf",             [" golf"]),
        ("Field Hockey",     ["field hockey"]),
        ("Gymnastics",       ["gymnastics"]),
        ("Bowling",          ["bowling"]),
    ]:
        if any(k in t for k in kws):
            return sport
    return "General/Multiple Sports"

NJ_SCHOOLS = sorted([
    "Hackettstown","Cedar Creek","Watchung Hills","Medford Tech","Wayne Valley",
    "Cedar Grove","South River","Old Bridge","Monroe","Immaculata","Mainland",
    "Egg Harbor Township","Clearview","Steinert","Allentown","Hightstown",
    "Lawrence","St. Joseph","South Brunswick","Passaic Valley","Don Bosco",
    "Seton Hall Prep","Bergen Catholic","Ridgewood","Northern Highlands",
    "Montclair","West Orange","Livingston","Summit","Westfield","Scotch Plains",
    "Rahway","Elizabeth","Linden","Union","Cranford","Millburn","Chatham",
    "Madison","Morristown","Randolph","Parsippany","Morris Hills","Mountain Lakes",
    "Roxbury","Mount Olive","West Morris","Pequannock","Hanover Park","Mendham",
    "Bernards","Somerville","Bound Brook","Manville","South Plainfield",
    "Edison","Piscataway","New Brunswick","East Brunswick","North Brunswick",
    "J.P. Stevens","Metuchen","Woodbridge","Perth Amboy","Sayreville","Freehold",
    "Howell","Jackson","Toms River","Southern","Ocean Township","Asbury Park",
    "Red Bank","Rumson","Manasquan","Point Pleasant","Brick","Lakewood","Lacey",
    "Long Branch","Matawan","Kingsway","Deptford","Gloucester","Haddonfield",
    "Cherry Hill","Moorestown","Shawnee","Lenape","Cherokee","Seneca",
    "Rancocas Valley","Pennsauken","Camden","Collingswood","Haddon Township",
    "Triton","Absegami","Atlantic City","Holy Spirit","St. Augustine",
    "Bishop Eustace","Paul VI","Gloucester Catholic","St. John Vianney",
    "Red Bank Catholic","Donovan Catholic","Pope John","Morris Catholic",
    "Delbarton","Pingry","Westwood","River Dell","Hackensack","Paramus Catholic",
    "Don Bosco Prep","St. Peter's Prep","Wood-Ridge","Lyndhurst","Rutherford",
    "Garfield","Clifton","Passaic","West Milford","Jefferson","Sparta","Vernon",
    "High Point","Newton","Sussex Tech","Kittatinny","Lenape Valley",
], key=len, reverse=True)

def extract_school(title):
    for school in NJ_SCHOOLS:
        if school.lower() in title.lower():
            return school
    return ""

# ── Load tracking ────────────────────────────────────────────────────────────
def load_seen_urls():
    if TRACKING_FILE.exists():
        with open(TRACKING_FILE) as f:
            return set(json.load(f))
    return set()

def save_seen_urls(seen):
    with open(TRACKING_FILE, "w") as f:
        json.dump(sorted(seen), f)

# ── Fetch RSS (paginated) ────────────────────────────────────────────────────
def fetch_rss_articles():
    articles = []
    offset = 0
    while True:
        try:
            content = fetch(RSS_BASE.format(offset))
            root = ET.fromstring(content)
            channel = root.find("channel")
            items = channel.findall("item")
            if not items:
                break
            oldest_in_batch = None
            for item in items:
                title   = item.findtext("title","").strip()
                link    = item.findtext("link","").strip()
                pub_raw = item.findtext("pubDate","").strip()
                desc    = item.findtext("description","").strip()[:300]
                author  = item.findtext("{http://purl.org/dc/elements/1.1/}creator","").strip()
                try:
                    dt = datetime.strptime(pub_raw, "%a, %d %b %Y %H:%M:%S %z")
                except:
                    continue
                oldest_in_batch = dt
                if START_DATE <= dt <= END_DATE:
                    articles.append({
                        "title":   title,
                        "date":    dt.strftime("%Y-%m-%d"),
                        "author":  author,
                        "link":    link,
                        "summary": desc,
                        "sport":   parse_sport(title),
                        "school":  extract_school(title),
                        "date_added": datetime.now().strftime("%Y-%m-%d"),
                    })
            # Stop if oldest article in batch is before our start date
            if oldest_in_batch and oldest_in_batch < START_DATE:
                break
            offset += 50
            time.sleep(0.3)
        except Exception as e:
            log(f"RSS fetch error at offset {offset}: {e}")
            break
    return articles

# ── Append to Excel ──────────────────────────────────────────────────────────
def append_to_excel(new_articles):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data_font = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", start_color="EBF3FB")
    border    = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC")
    )
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    current_row = ws.max_row + 1

    for art in new_articles:
        fill = alt_fill if current_row % 2 == 0 else None
        row_data = [
            art["title"], art["date"], art["school"], art["sport"],
            art["link"], art["summary"], art["author"], art["date_added"],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = data_font
            cell.alignment = left_wrap
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    wb.save(EXCEL_FILE)
    log(f"Appended {len(new_articles)} new articles. Total rows: {current_row - 2}")

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    log("=== NJ HS Sports Scraper starting ===")
    seen_urls = load_seen_urls()
    log(f"Loaded {len(seen_urls)} previously seen URLs")

    rss_articles = fetch_rss_articles()
    log(f"Fetched {len(rss_articles)} articles from RSS feed")

    new_articles = [a for a in rss_articles if a["link"] not in seen_urls]
    log(f"New articles (not yet in spreadsheet): {len(new_articles)}")

    if new_articles:
        # Sort oldest-first so spreadsheet stays chronological
        new_articles.sort(key=lambda x: x["date"])
        append_to_excel(new_articles)
        for a in new_articles:
            seen_urls.add(a["link"])
        save_seen_urls(seen_urls)
        log(f"Done. Added {len(new_articles)} articles to {EXCEL_FILE.name}")
    else:
        log("No new articles found — spreadsheet is up to date.")

    log("=== Scraper complete ===\n")

if __name__ == "__main__":
    main()
